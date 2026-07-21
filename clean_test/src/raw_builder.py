from __future__ import annotations

import csv
import logging
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import torch
from openpyxl import load_workbook
from scipy.interpolate import PchipInterpolator

from .utils import csv_safe, safe_torch_load, validate_archive_size, write_json

LOGGER = logging.getLogger(__name__)
VARIABLES = ("Te", "ne", "Ti")
MAX_XLSX_BYTES = 64 * 1024**2
MAX_XLSX_UNCOMPRESSED_BYTES = 256 * 1024**2
MAX_SHEET_ROWS = 100_000
MAX_PROFILE_CSV_BYTES = 64 * 1024**2
MAX_PROFILE_ROWS = 1_000_000
INTERPOLATION_METHOD = "pchip"
BOUNDARY_ATOL = 1e-7
MIN_X_SPACING = 1e-8


@dataclass
class RawProfile:
    """One metadata row plus its explicit CSV measurements."""

    profile_id: str
    paper_id: str
    case_id: str
    variable: str
    coord_name: str
    unit_raw: str
    status: str
    csv_path: Path
    x: np.ndarray
    y: np.ndarray


def _clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _read_sheet_table(path: Path, sheet_name: str, required_column: str) -> list[dict[str, Any]]:
    validate_archive_size(
        path,
        max_file_bytes=MAX_XLSX_BYTES,
        max_uncompressed_bytes=MAX_XLSX_UNCOMPRESSED_BYTES,
        max_members=5_000,
    )
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"{path.name} 缺少工作表 {sheet_name}")
        headers: list[str] | None = None
        result: list[dict[str, Any]] = []
        for index, row in enumerate(workbook[sheet_name].iter_rows(values_only=True)):
            if index >= MAX_SHEET_ROWS:
                raise ValueError(f"{path.name}/{sheet_name} 超过 {MAX_SHEET_ROWS} 行限制")
            if headers is None:
                if index >= 15:
                    break
                candidate = [_clean_text(value) for value in row]
                if required_column in set(candidate):
                    headers = candidate
                continue
            record = {
                headers[i]: value
                for i, value in enumerate(row)
                if i < len(headers) and headers[i]
            }
            if _clean_text(record.get(required_column)):
                result.append(record)
        if headers is None:
            raise ValueError(f"{path.name}/{sheet_name} 前 15 行找不到字段 {required_column}")
        return result
    finally:
        workbook.close()


def _numeric(value: Any) -> float | None:
    if isinstance(value, (int, float)) and np.isfinite(value):
        return float(value)
    text = _clean_text(value)
    try:
        number = float(text)
    except ValueError:
        return None
    return number if np.isfinite(number) else None


def _read_xy(path: Path) -> tuple[np.ndarray, np.ndarray, list[str]]:
    if path.is_symlink():
        raise ValueError(f"拒绝读取符号链接 CSV: {path}")
    if path.stat().st_size > MAX_PROFILE_CSV_BYTES:
        raise ValueError(f"CSV 超过 {MAX_PROFILE_CSV_BYTES} bytes 限制: {path}")
    rows: list[tuple[float, float]] = []
    numeric_started = False
    for line_no, raw_line in enumerate(
        path.read_text(encoding="utf-8-sig").splitlines(), start=1
    ):
        if line_no > MAX_PROFILE_ROWS:
            raise ValueError(f"CSV 超过 {MAX_PROFILE_ROWS} 行限制: {path}")
        if len(raw_line) > 1024 * 1024:
            raise ValueError(f"{path} 第 {line_no} 行超过 1 MiB")
        parts = [part for part in re.split(r"[,;\s]+", raw_line.strip()) if part]
        if len(parts) < 2:
            if numeric_started and parts:
                raise ValueError(f"{path} 第 {line_no} 行不是有效的 x,y 数据")
            continue
        try:
            rows.append((float(parts[0]), float(parts[1])))
        except ValueError:
            if numeric_started:
                raise ValueError(f"{path} 第 {line_no} 行不是有效的 x,y 数值")
            continue
        numeric_started = True
    if len(rows) < 2:
        raise ValueError(f"{path} 至少需要两行数值 x,y。")
    array = np.asarray(rows, dtype=np.float64)
    finite = np.isfinite(array).all(axis=1)
    if not finite.all():
        raise ValueError(f"{path} 包含 NaN 或 Inf。")
    order = np.argsort(array[:, 0], kind="stable")
    x, y = array[order, 0], array[order, 1]
    warnings: list[str] = []
    unique_x, inverse, counts = np.unique(x, return_inverse=True, return_counts=True)
    if np.any(counts > 1):
        averaged = np.zeros_like(unique_x)
        np.add.at(averaged, inverse, y)
        y = averaged / counts
        x = unique_x
        warnings.append(f"{path}: 存在重复 x，已按相同 x 显式取均值后插值。")
    if len(x) < 2 or not np.all(np.diff(x) > 0):
        raise ValueError(f"{path} 的 x 无法整理为严格递增。")
    minimum_spacing = float(np.diff(x).min())
    if minimum_spacing <= MIN_X_SPACING:
        raise ValueError(
            f"{path} 的相邻 x 最小间距 {minimum_spacing:.3g} 不大于安全阈值 "
            f"{MIN_X_SPACING:.3g}；拒绝进入 PCHIP，避免病态斜率。"
        )
    return x, y, warnings


def _canonical_unit(unit: str) -> str:
    return (
        unit.lower()
        .replace(" ", "")
        .replace("³", "3")
        .replace("−", "-")
        .replace("×", "x")
        .replace("^", "")
    )


def _convert_units(variable: str, y: np.ndarray, unit_raw: str) -> tuple[np.ndarray, str]:
    canonical = _canonical_unit(unit_raw)
    if variable in {"Te", "Ti"}:
        if canonical == "kev":
            return y, f"{unit_raw} -> keV (factor=1)"
        if canonical == "ev":
            return y / 1000.0, f"{unit_raw} -> keV (factor=0.001)"
    if variable == "ne":
        if canonical in {"m-3", "m-3(si)"}:
            return y / 1e19, f"{unit_raw} -> 10^19 m^-3 (factor=1e-19)"
        if canonical in {"1019m-3", "1e19m-3"}:
            return y, f"{unit_raw} -> 10^19 m^-3 (factor=1)"
        if canonical in {"1020m-3", "1e20m-3"}:
            return y * 10.0, f"{unit_raw} -> 10^19 m^-3 (factor=10)"
    raise ValueError(f"变量 {variable} 的单位无法识别: {unit_raw!r}；不会根据数值大小猜测。")


def _suffix(profile_id: str, variable: str) -> str:
    tokens = [token.strip().lower() for token in profile_id.split("__")]
    for index, token in enumerate(tokens):
        if token == variable.lower():
            return "__".join(tokens[index + 1 :])
    return tokens[-1] if tokens else ""


def _locate_csv(
    dataset_root: Path,
    profile: dict[str, Any],
    all_csv: list[Path],
) -> tuple[Path | None, list[str]]:
    profile_id = _clean_text(profile.get("profile_id"))
    warnings: list[str] = []
    exact = [path for path in all_csv if path.stem.strip() == profile_id]
    if len(exact) == 1:
        return exact[0], warnings
    if len(exact) > 1:
        warnings.append(
            f"{profile_id}: 存在多个同名 CSV，无法唯一确定: "
            f"{[str(path.relative_to(dataset_root)) for path in exact]}"
        )
        return None, warnings
    for key in ("target_points_csv_relpath", "points_file_relpath"):
        rel = _clean_text(profile.get(key))
        if not rel:
            continue
        candidate = (dataset_root / Path(rel.replace("\\", "/"))).resolve()
        try:
            candidate.relative_to(dataset_root.resolve())
        except ValueError:
            warnings.append(f"{profile_id}: 元数据 CSV 路径越出 raw 根目录，已拒绝: {rel}")
            continue
        if candidate.exists() and candidate.suffix.lower() == ".csv":
            if candidate.stem.strip() != profile_id:
                warnings.append(
                    f"{profile_id}: 元数据路径指向 {candidate.name}，与 profile_id 不一致；已拒绝猜测。"
                )
                return None, warnings
            return candidate, warnings
    warnings.append(f"{profile_id}: 没有文件名与 profile_id 精确一致的 CSV；已拒绝模糊匹配。")
    return None, warnings


def _match_reference(target: RawProfile, candidates: list[RawProfile]) -> RawProfile:
    if target.variable == candidates[0].variable and target in candidates:
        return target
    if len(candidates) == 1:
        return candidates[0]
    target_suffix = _suffix(target.profile_id, target.variable)
    exact = [candidate for candidate in candidates if _suffix(candidate.profile_id, candidate.variable) == target_suffix]
    if len(exact) == 1:
        return exact[0]
    raise ValueError(
        f"{target.profile_id}: 多条参考剖面中没有唯一的精确后缀匹配；"
        f"候选为 {[value.profile_id for value in candidates]}。拒绝启发式匹配。"
    )


def _evaluate_pchip(profile: RawProfile, points: np.ndarray) -> np.ndarray:
    """Evaluate PCHIP inside coverage and reject non-finite or overshooting results."""
    points = np.asarray(points, dtype=np.float64)
    spacing = np.diff(profile.x)
    if len(profile.x) < 2 or np.any(spacing <= MIN_X_SPACING):
        raise ValueError(
            f"{profile.profile_id}: PCHIP 要求至少两个严格递增且间距大于 "
            f"{MIN_X_SPACING:.3g} 的原始 x。"
        )
    interpolator = PchipInterpolator(profile.x, profile.y, extrapolate=False)
    values = np.asarray(interpolator(points), dtype=np.float64)
    if not np.isfinite(values).all():
        raise ValueError(f"{profile.profile_id}: PCHIP 产生 NaN/Inf 或发生了范围外求值。")

    interval = np.searchsorted(profile.x, points, side="right") - 1
    interval = np.clip(interval, 0, len(profile.x) - 2)
    local_low = np.minimum(profile.y[interval], profile.y[interval + 1])
    local_high = np.maximum(profile.y[interval], profile.y[interval + 1])
    value_scale = max(1.0, float(np.max(np.abs(profile.y))))
    value_tolerance = 1e-10 * value_scale
    if np.any(values < local_low - value_tolerance) or np.any(
        values > local_high + value_tolerance
    ):
        raise ValueError(f"{profile.profile_id}: PCHIP 结果越过相邻原始点范围，已拒绝。")
    return values


def _interpolate_value(profile: RawProfile, rho: float) -> float:
    if rho < profile.x[0] - BOUNDARY_ATOL or rho > profile.x[-1] + BOUNDARY_ATOL:
        raise ValueError(
            f"{profile.profile_id}: 原始坐标范围 [{profile.x[0]:.4g},{profile.x[-1]:.4g}] 不覆盖 rho={rho}。"
        )
    evaluation_point = np.asarray([np.clip(rho, profile.x[0], profile.x[-1])])
    return float(_evaluate_pchip(profile, evaluation_point)[0])


def _resample(profile: RawProfile, grid: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    grid = np.asarray(grid, dtype=np.float64)
    mask = (grid >= profile.x[0] - BOUNDARY_ATOL) & (
        grid <= profile.x[-1] + BOUNDARY_ATOL
    )
    target = np.zeros_like(grid, dtype=np.float32)
    evaluation_points = np.clip(grid[mask], profile.x[0], profile.x[-1])
    target[mask] = _evaluate_pchip(profile, evaluation_points).astype(np.float32)
    return target, mask.astype(np.float32)


def build_clean_datasets(
    dataset_root: Path,
    output_dir: Path,
    variables: Iterable[str] = VARIABLES,
    min_valid_points: int = 4,
    strict: bool = False,
    dataset_role: str | None = None,
) -> dict[str, Any]:
    """Build variable-specific .pt datasets from heterogeneous metadata workbooks."""
    workbook_dir = dataset_root / "database_by_paper"
    data_dir = dataset_root / "data"
    if not workbook_dir.exists() or not data_dir.exists():
        raise FileNotFoundError("dataset_root 必须包含 database_by_paper/ 和 data/。")
    output_dir.mkdir(parents=True, exist_ok=True)
    all_csv: list[Path] = []
    raw_root = dataset_root.resolve()
    for path in sorted(data_dir.rglob("*.csv")):
        if path.is_symlink():
            LOGGER.warning("忽略符号链接 CSV: %s", path)
            continue
        try:
            path.resolve().relative_to(raw_root)
        except ValueError:
            LOGGER.warning("忽略越出 raw 根目录的 CSV: %s", path)
            continue
        all_csv.append(path)
    warnings: list[str] = []
    exclusions: list[dict[str, str]] = []
    unit_log: list[dict[str, str]] = []
    cases: dict[str, dict[str, Any]] = {}
    duplicate_case_ids: set[str] = set()
    raw_profiles: list[RawProfile] = []

    for workbook_path in sorted(workbook_dir.glob("*.xlsx")):
        try:
            case_rows = _read_sheet_table(workbook_path, "02_CASES", "case_id")
            profile_rows = _read_sheet_table(workbook_path, "03_PROFILES", "profile_id")
        except Exception as exc:
            if strict:
                raise
            exclusions.append({"entity": workbook_path.name, "reason": str(exc)})
            continue
        for row in case_rows:
            case_id = _clean_text(row.get("case_id"))
            paper_id = _clean_text(row.get("paper_id"))
            delta = _numeric(row.get("delta", row.get("delta_avg")))
            if case_id in cases:
                duplicate_case_ids.add(case_id)
                exclusions.append({"entity": case_id, "reason": "case_id 在多个元数据行中重复"})
                continue
            cases[case_id] = {"paper_id": paper_id, "delta": delta, "source": workbook_path.name}
        for row in profile_rows:
            profile_id = _clean_text(row.get("profile_id"))
            variable = _clean_text(row.get("variable"))
            if variable not in VARIABLES:
                continue
            if _clean_text(row.get("data_origin")).lower() != "experimental":
                exclusions.append({"entity": profile_id, "reason": "data_origin 不是 experimental"})
                continue
            status = _clean_text(row.get("status")).lower()
            if status == "planned":
                warnings.append(f"{profile_id}: status=planned，但 CSV 已存在时允许进入当前试验数据集。")
            case_id = _clean_text(row.get("case_id"))
            profile_paper_id = _clean_text(row.get("paper_id"))
            case_paper_id = _clean_text(cases.get(case_id, {}).get("paper_id"))
            if profile_paper_id and case_paper_id and profile_paper_id != case_paper_id:
                exclusions.append(
                    {
                        "entity": profile_id,
                        "reason": (
                            f"03_PROFILES.paper_id={profile_paper_id!r} 与 "
                            f"02_CASES.paper_id={case_paper_id!r} 不一致"
                        ),
                    }
                )
                continue
            path, path_warnings = _locate_csv(dataset_root, row, all_csv)
            warnings.extend(path_warnings)
            if path is None:
                exclusions.append({"entity": profile_id, "reason": "找不到唯一 CSV"})
                continue
            try:
                x, y, xy_warnings = _read_xy(path)
                warnings.extend(xy_warnings)
                converted, conversion = _convert_units(variable, y, _clean_text(row.get("y_unit_raw")))
            except Exception as exc:
                if strict:
                    raise
                exclusions.append({"entity": profile_id, "reason": str(exc)})
                continue
            unit_log.append({"profile_id": profile_id, "conversion": conversion})
            raw_profiles.append(
                RawProfile(
                    profile_id=profile_id,
                    paper_id=profile_paper_id,
                    case_id=case_id,
                    variable=variable,
                    coord_name=_clean_text(row.get("coord_name")),
                    unit_raw=_clean_text(row.get("y_unit_raw")),
                    status=status,
                    csv_path=path,
                    x=x,
                    y=converted,
                )
            )

    profile_counts: dict[str, int] = defaultdict(int)
    for profile in raw_profiles:
        profile_counts[profile.profile_id] += 1
    duplicate_profile_ids = {key for key, count in profile_counts.items() if count > 1}
    for profile_id in sorted(duplicate_profile_ids):
        exclusions.append({"entity": profile_id, "reason": "profile_id 在多个元数据行中重复"})

    by_case: dict[str, dict[str, list[RawProfile]]] = defaultdict(lambda: defaultdict(list))
    for profile in raw_profiles:
        if profile.profile_id in duplicate_profile_ids:
            continue
        by_case[profile.case_id][profile.variable].append(profile)
    # Interpolate on an exact float64 grid; serialize tensors as float32 below.
    grid = np.linspace(0.8, 1.0, 64, dtype=np.float64)
    built: dict[str, dict[str, list[Any]]] = {
        variable: {
            "case_ids": [],
            "paper_ids": [],
            "profile_ids": [],
            "branch_input": [],
            "targets": [],
            "valid_mask": [],
            "coord_names": [],
        }
        for variable in variables
    }
    for case_id, profiles_by_variable in sorted(by_case.items()):
        if case_id in duplicate_case_ids:
            continue
        case = cases.get(case_id)
        if case is None:
            exclusions.append({"entity": case_id, "reason": "profile 引用的 case_id 不在 02_CASES"})
            continue
        if case["delta"] is None:
            exclusions.append({"entity": case_id, "reason": "delta/delta_avg 缺失或不是明确数值"})
            continue
        if float(case["delta"]) >= 0:
            exclusions.append({"entity": case_id, "reason": "不是明确的负三角形变(delta<0)"})
            continue
        if not profiles_by_variable.get("Te") or not profiles_by_variable.get("ne"):
            exclusions.append({"entity": case_id, "reason": "缺少 Te 或 ne，无法构造 [delta,T0,n0]"})
            continue
        for variable in variables:
            for target_profile in profiles_by_variable.get(variable, []):
                try:
                    te_ref = _match_reference(target_profile, profiles_by_variable["Te"])
                    ne_ref = _match_reference(target_profile, profiles_by_variable["ne"])
                    t0 = _interpolate_value(te_ref, 0.8)
                    n0 = _interpolate_value(ne_ref, 0.8)
                    target, mask = _resample(target_profile, grid)
                    if int(mask.sum()) < min_valid_points:
                        raise ValueError(
                            f"rho=[0.8,1.0] 上仅 {int(mask.sum())} 个有效网格点，少于 {min_valid_points}。"
                        )
                    coord_set = {target_profile.coord_name, te_ref.coord_name, ne_ref.coord_name}
                    if len(coord_set) > 1:
                        warnings.append(
                            f"{target_profile.profile_id}: target/Te/ne 坐标定义不同 {sorted(coord_set)}。"
                        )
                    record = built[variable]
                    record["case_ids"].append(case_id)
                    record["paper_ids"].append(case["paper_id"] or target_profile.paper_id)
                    record["profile_ids"].append(target_profile.profile_id)
                    record["branch_input"].append([float(case["delta"]), t0, n0])
                    record["targets"].append(target)
                    record["valid_mask"].append(mask)
                    record["coord_names"].append(target_profile.coord_name)
                except Exception as exc:
                    if strict:
                        raise
                    exclusions.append({"entity": target_profile.profile_id, "reason": str(exc)})

    files: dict[str, str] = {}
    for variable, record in built.items():
        if not record["case_ids"]:
            warnings.append(f"{variable}: 没有可用样本，未生成数据集。")
            continue
        coord_names = sorted(set(record["coord_names"]))
        if len(coord_names) > 1:
            warnings.append(
                f"{variable}: 混合使用归一化坐标 {coord_names}；它们并非严格等价，后续应统一坐标。"
            )
        payload = {
            "case_ids": record["case_ids"],
            "paper_ids": record["paper_ids"],
            "profile_ids": record["profile_ids"],
            "branch_input": torch.tensor(record["branch_input"], dtype=torch.float32),
            "trunk_x": torch.tensor(grid, dtype=torch.float32),
            "targets": torch.from_numpy(np.stack(record["targets"])),
            "valid_mask": torch.from_numpy(np.stack(record["valid_mask"])),
            "variable": variable,
            "metadata": {
                "source_root": str(dataset_root.resolve()),
                "data_stage": "clean",
                "dataset_role": dataset_role,
                "coord_names": coord_names,
                "units": "keV" if variable in {"Te", "Ti"} else "10^19 m^-3",
                "interpolation_method": INTERPOLATION_METHOD,
                "interpolation_extrapolate": False,
                "interpolation_boundary_atol": BOUNDARY_ATOL,
                "minimum_raw_x_spacing": MIN_X_SPACING,
                "note": "branch_scaled/scaler_mean/scaler_std intentionally absent; fit on train only.",
            },
        }
        path = output_dir / f"deeponet_dataset_{variable}.pt"
        torch.save(payload, path)
        files[variable] = str(path)
    report = {
        "status": "ok" if files else "error",
        "dataset_role": dataset_role,
        "raw_source": str(dataset_root.resolve()),
        "clean_output": str(output_dir.resolve()),
        "generated_files": files,
        "sample_counts": {variable: len(record["case_ids"]) for variable, record in built.items()},
        "warnings": sorted(set(warnings)),
        "exclusions": exclusions,
        "unit_conversions": unit_log,
        "interpolation": {
            "method": INTERPOLATION_METHOD,
            "extrapolate": False,
            "boundary_atol": BOUNDARY_ATOL,
            "minimum_raw_x_spacing": MIN_X_SPACING,
        },
        "assumptions": [
            "不同论文给出的归一化径向坐标暂映射到共同 rho 网格；报告会列出混合坐标风险。",
            "使用 shape-preserving PCHIP，只在原始 x 覆盖范围内插值，不做径向外推。",
            "仅将距离原始端点不超过 1e-7 的网格点贴到端点，用于吸收浮点表示误差。",
            "T0 和 n0 是匹配 Te/ne 剖面在 rho=0.8 的 PCHIP 插值值。",
            "status=planned 且 CSV 实际存在时纳入，并保留警告。",
        ],
    }
    write_json(report, output_dir / "build_report.json")
    with (output_dir / "unit_conversion_log.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["profile_id", "conversion"])
        writer.writeheader()
        writer.writerows(
            {key: csv_safe(value) for key, value in row.items()} for row in unit_log
        )
    if not files:
        raise ValueError("当前原始数据没有构建出任何可训练样本，详见 build_report.json。")
    return report


def write_role_manifest(dataset_path: Path, output_path: Path, role: str) -> None:
    """Write all unique cases as one manually declared train/val/test role."""
    if role not in {"train", "val", "test"}:
        raise ValueError(f"Unsupported dataset role: {role}")
    payload = safe_torch_load(dataset_path, map_location="cpu")
    case_ids = sorted(set(str(value) for value in payload["case_ids"]))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["case_id", "split"])
        writer.writerows((csv_safe(case_id), role) for case_id in case_ids)


def write_train_only_manifest(dataset_path: Path, output_path: Path) -> None:
    """Backward-compatible wrapper for older commands."""
    write_role_manifest(dataset_path, output_path, "train")
